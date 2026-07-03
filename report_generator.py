from datetime import datetime
from pathlib import Path
from typing import List, Dict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from email_clients import EmailMessage


class ReportGenerator:
    """Generate Excel reports from email data"""
    
    def __init__(self, output_dir: Path = Path('./reports')):
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(exist_ok=True)
        
    def generate_report(self, emails_by_category: Dict[str, List[EmailMessage]]) -> Path:
        """Generate Excel report with multiple sheets"""
        try:
            workbook = Workbook()
            workbook.remove(workbook.active)  # Remove default sheet
            
            # Create sheets for each category
            for category, emails in emails_by_category.items():
                if emails:
                    self._create_sheet(workbook, category, emails)
            
            # Create summary sheet
            self._create_summary_sheet(workbook, emails_by_category)
            
            # Save workbook
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"email_report_{timestamp}.xlsx"
            filepath = self.output_dir / filename
            
            workbook.save(filepath)
            print(f"✓ Report generated: {filepath}")
            return filepath
            
        except Exception as e:
            print(f"✗ Error generating report: {e}")
            raise

    def _create_sheet(self, workbook: Workbook, sheet_name: str, emails: List[EmailMessage]) -> None:
        """Create a worksheet with email data"""
        sheet = workbook.create_sheet(title=sheet_name[:31])  # Excel max 31 chars
        
        # Define headers
        headers = ['Remitente', 'Asunto', 'Fecha', 'Cuerpo (primeros 500 caracteres)', 'Fuente']
        
        # Add headers with styling
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for col_num, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # Add data rows
        for row_num, email in enumerate(emails, 2):
            data = [
                email.sender,
                email.subject,
                email.date.strftime('%Y-%m-%d %H:%M:%S') if email.date else 'N/A',
                email.body[:500],
                email.source.upper()
            ]
            
            for col_num, value in enumerate(data, 1):
                cell = sheet.cell(row=row_num, column=col_num)
                cell.value = value
                cell.border = border
                cell.alignment = Alignment(vertical='top', wrap_text=True)
        
        # Adjust column widths
        sheet.column_dimensions['A'].width = 25
        sheet.column_dimensions['B'].width = 40
        sheet.column_dimensions['C'].width = 18
        sheet.column_dimensions['D'].width = 50
        sheet.column_dimensions['E'].width = 12
        
        # Set row height for header
        sheet.row_dimensions[1].height = 25

    def _create_summary_sheet(self, workbook: Workbook, emails_by_category: Dict[str, List[EmailMessage]]) -> None:
        """Create a summary sheet with statistics"""
        sheet = workbook.create_sheet(title='Resumen', index=0)
        
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF', size=14)
        data_font = Font(size=12)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Title
        sheet['A1'] = 'RESUMEN DE CORREOS'
        sheet['A1'].font = Font(bold=True, size=16)
        sheet.merge_cells('A1:C1')
        
        # Timestamp
        sheet['A2'] = f"Generado: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        sheet['A2'].font = data_font
        sheet.merge_cells('A2:C2')
        
        # Statistics
        row = 4
        sheet[f'A{row}'] = 'Categoría'
        sheet[f'B{row}'] = 'Total de Correos'
        sheet[f'C{row}'] = 'Fuentes'
        
        for col in ['A', 'B', 'C']:
            cell = sheet[f'{col}{row}']
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = Alignment(horizontal='center')
        
        row += 1
        total_emails = 0
        for category, emails in emails_by_category.items():
            count = len(emails)
            sources = ', '.join(sorted(set(e.source for e in emails)))
            
            sheet[f'A{row}'] = category
            sheet[f'B{row}'] = count
            sheet[f'C{row}'] = sources
            
            for col in ['A', 'B', 'C']:
                cell = sheet[f'{col}{row}']
                cell.border = border
                cell.alignment = Alignment(horizontal='left')
            
            total_emails += count
            row += 1
        
        # Total row
        sheet[f'A{row}'] = 'TOTAL'
        sheet[f'B{row}'] = total_emails
        sheet[f'A{row}'].font = Font(bold=True)
        sheet[f'B{row}'].font = Font(bold=True)
        
        for col in ['A', 'B', 'C']:
            cell = sheet[f'{col}{row}']
            cell.fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
            cell.border = border
        
        # Set column widths
        sheet.column_dimensions['A'].width = 25
        sheet.column_dimensions['B'].width = 18
        sheet.column_dimensions['C'].width = 30
